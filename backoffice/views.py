from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Count, Sum, F, ExpressionWrapper, DecimalField, Value, Q, Subquery, OuterRef, Window
from django.db.models.functions import Coalesce, TruncMonth, TruncDate, RowNumber

from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.utils import timezone

from urllib.parse import urlencode
from decimal import Decimal
import csv
import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors

from accounts.decorators import admin_required, content_staff_required
from orders.models import Order, OrderItem, PartnerPayout
from partners.models import PartnerProfile
from accounts.models import User, ClientStructure
from catalog.models import Product, Category, KitComponent,ProductRating
from cms.models import Page

class NumberedCanvas(canvas.Canvas):
    """
    Canvas personalizzato che aggiunge in automatico:
    - footer con "Preparato da B2B Turismo – dd/mm/yyyy"
    - numero pagina "Pagina X di Y"
    su ogni pagina del PDF.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        # Salva lo stato corrente della pagina prima di passare alla successiva
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        # Alla fine, disegna footer e numero pagina su tutte le pagine
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page_footer(page_count)
            super().showPage()
        super().save()

    def _draw_page_footer(self, page_count: int):
        width, height = self._pagesize
        prepared_on = timezone.now().strftime("%d/%m/%Y")

        self.setFont("Helvetica", 8)
        self.setFillColor(colors.grey)

        # Riga "Preparato da..."
        self.drawCentredString(
            width / 2.0,
            15 * mm,
            f"Preparato da B2B Turismo – {prepared_on}",
        )

        # Riga "Pagina X di Y"
        self.drawCentredString(
            width / 2.0,
            10 * mm,
            f"Pagina {self._pageNumber} di {page_count}",
        )

        self.setFillColor(colors.black)

# -------------------------------
# DASHBOARD ADMIN
# -------------------------------
@admin_required
def dashboard(request):

    # KPI: ordini per stato
    orders_by_status = (
        Order.objects.values("status")
        .annotate(total=Count("id"))
        .order_by("status")
    )

    # Fatturato totale (escludendo annullati)
    total_revenue = (
        Order.objects.exclude(status="cancelled")
        .aggregate(total=Sum("total"))
    )["total"] or 0

    # Top partner per fatturato
    partner_revenue = (
        OrderItem.objects.exclude(order__status="cancelled")
        .values("partner__id", "partner__company_name")
        .annotate(total=Sum("total_price"))
        .order_by("-total")[:5]
    )

    # Ordini critici (pending o processing)
    critical_orders = (
        Order.objects.filter(status__in=["pending_payment", "processing"])
        .order_by("-created_at")[:10]
    )

    context = {
        "orders_by_status": orders_by_status,
        "total_revenue": total_revenue,
        "partner_revenue": partner_revenue,
        "critical_orders": critical_orders,
    }
    return render(request, "backoffice/dashboard.html", context)


@admin_required
def dashboard_live_stats(request):
    """
    Endpoint JSON per la dashboard LIVE:
    - nuovo ordine
    - nuova commissione maturata
    - nuovo payout generato
    """

    # Ordini (escludiamo i cancellati)
    order_qs = Order.objects.exclude(status=Order.STATUS_CANCELLED)
    order_count = order_qs.count()
    latest_order = order_qs.order_by("-created_at").first()

    # Righe con commissioni maturate
    commission_qs = OrderItem.objects.exclude(
        order__status=Order.STATUS_CANCELLED
    ).filter(
        commission_amount__gt=0
    )
    commission_count = commission_qs.count()
    latest_commission_item = commission_qs.order_by("-order__created_at", "-id").first()

    # Payout partner
    payout_qs = PartnerPayout.objects.all()
    payout_count = payout_qs.count()
    latest_payout = payout_qs.order_by("-created_at").first()

    data = {
        "order_count": order_count,
        "last_order_id": latest_order.id if latest_order else None,
        "last_order_created_at": latest_order.created_at.isoformat() if latest_order else None,

        "commission_count": commission_count,
        "last_commission_id": latest_commission_item.id if latest_commission_item else None,
        "last_commission_order_id": latest_commission_item.order.id if latest_commission_item else None,
        "last_commission_created_at": latest_commission_item.order.created_at.isoformat() if latest_commission_item else None,

        "payout_count": payout_count,
        "last_payout_id": latest_payout.id if latest_payout else None,
        "last_payout_created_at": latest_payout.created_at.isoformat() if latest_payout else None,
    }
    return JsonResponse(data)


# -------------------------------
# LISTA ORDINI
# -------------------------------
@admin_required
def order_list(request):

    qs = Order.objects.select_related("client", "structure").all().order_by("-created_at")

    # Filtri
    status = request.GET.get("status")
    client_email = request.GET.get("client")
    partner_id = request.GET.get("partner")
    structure_id = request.GET.get("structure")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if status:
        qs = qs.filter(status=status)

    if client_email:
        qs = qs.filter(client__email__icontains=client_email)

    if partner_id:
        qs = qs.filter(items__partner__id=partner_id).distinct()
        
    if structure_id:
        qs = qs.filter(structure_id=structure_id)

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)

    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    partners = PartnerProfile.objects.filter(is_active=True)

    context = {
        "orders": qs,
        "partners": partners,
        "selected_status": status or "",
        "selected_partner": int(partner_id) if partner_id else None,
        "client_email": client_email or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
    }
    return render(request, "backoffice/order_list.html", context)


# -------------------------------
# DETTAGLIO ORDINE
# -------------------------------
@admin_required
def order_detail(request, order_id):

    order = get_object_or_404(
        Order.objects.select_related("client", "structure"),
        id=order_id
    )

    items = (
        OrderItem.objects
        .select_related("product", "partner")
        .filter(order=order)
    )

    # ✅ LOCK: se esiste almeno una riga liquidata o legata a payout PAID
    has_paid_payout = (
        items.filter(is_liquidated=True).exists()
        or items.filter(payout__status=PartnerPayout.STATUS_PAID).exists()
    )

    # ranking stati
    status_rank = {code: idx for idx, (code, _) in enumerate(Order.STATUS_CHOICES)}

    # stati disabilitati lato UI
    disabled_statuses = set()
    if has_paid_payout:
        cur = order.status

        if hasattr(Order, "STATUS_CANCELLED"):
            disabled_statuses.add(Order.STATUS_CANCELLED)

        if cur in status_rank:
            cur_rank = status_rank[cur]
            for value, _label in Order.STATUS_CHOICES:
                if value in status_rank and status_rank[value] < cur_rank:
                    disabled_statuses.add(value)

    if request.method == "POST":

        new_status = request.POST.get("status")
        admin_notes = request.POST.get("admin_notes", "")

        valid_statuses = {choice[0] for choice in Order.STATUS_CHOICES}

        # le note admin sono sempre aggiornabili
        order.admin_notes = admin_notes

        if new_status in valid_statuses:
            # 🔒 HARD LOCK: blocca downgrade / cancel dopo payout
            if has_paid_payout and new_status != order.status:

                if hasattr(Order, "STATUS_CANCELLED") and new_status == Order.STATUS_CANCELLED:
                    messages.error(
                        request,
                        "Operazione non consentita: esiste già un payout pagato per questo ordine."
                    )
                    return redirect("backoffice:order_detail", order_id=order.id)

                if (
                    new_status in status_rank
                    and order.status in status_rank
                    and status_rank[new_status] < status_rank[order.status]
                ):
                    messages.error(
                        request,
                        "Operazione non consentita: non puoi retrocedere lo stato dopo il payout."
                    )
                    return redirect("backoffice:order_detail", order_id=order.id)

            order.status = new_status

        order.save()
        messages.success(request, "Ordine aggiornato correttamente.")
        return redirect("backoffice:order_detail", order_id=order.id)

    context = {
        "order": order,
        "items": items,
        "has_paid_payout": has_paid_payout,
        "disabled_statuses": disabled_statuses,
    }

    return render(request, "backoffice/order_detail.html", context)

# -------------------------------
# LISTA PARTNER
# -------------------------------
@admin_required
def partner_list(request):

    partners = (
        PartnerProfile.objects.select_related("user")
        .all()
        .order_by("company_name")
    )

    context = {
        "partners": partners,
    }
    return render(request, "backoffice/partner_list.html", context)


# -------------------------------
# LISTA CLIENTI
# -------------------------------
@admin_required
def client_list(request):

    clients = User.objects.filter(role="client").order_by("email")

    context = {
        "clients": clients,
    }
    return render(request, "backoffice/client_list.html", context)


# -------------------------------
# LISTA PRODOTTI
# -------------------------------
@admin_required
def product_list(request):

    products = (
        Product.objects.select_related("category", "supplier")
        .all()
        .order_by("name")
    )

    context = {
        "products": products,
    }
    return render(request, "backoffice/product_list.html", context)
    
# -------------------------------
# LISTA STRUTTURE CLIENTE
# -------------------------------

from accounts.models import ClientStructure  # assicurati che questo import ci sia in alto


@admin_required
def client_structure_list(request):

    qs = (
        ClientStructure.objects
        .select_related("owner")
        .all()
        .order_by("owner__email", "name")
    )

    # Filtri
    client_email = request.GET.get("client")
    structure_name = request.GET.get("name")

    if client_email:
        qs = qs.filter(owner__email__icontains=client_email)

    if structure_name:
        qs = qs.filter(name__icontains=structure_name)

    context = {
        "structures": qs,
        "client_email": client_email or "",
        "structure_name": structure_name or "",
    }
    return render(request, "backoffice/client_structure_list.html", context)
    

@admin_required
def client_structure_detail(request, pk):
    """
    Dettaglio di una struttura cliente, con riepilogo ordini collegati.
    """
    structure = get_object_or_404(
        ClientStructure.objects.select_related("owner"),
        pk=pk
    )

    # grazie al related_name "orders" visto nell'errore, possiamo usare structure.orders
    orders = structure.orders.select_related("client").all().order_by("-created_at")

    context = {
        "structure": structure,
        "orders": orders,
    }
    return render(request, "backoffice/client_structure_detail.html", context)
    
# -------------------------------
# LISTA CATEGORIE PRODOTTO
# -------------------------------
@admin_required
def category_list(request):
    qs = Category.objects.all().order_by("name")

    # Filtro per nome categoria
    name = request.GET.get("name")
    if name:
        qs = qs.filter(name__icontains=name)

    context = {
        "categories": qs,
        "name": name or "",
    }
    return render(request, "backoffice/category_list.html", context)
    
    
# -------------------------------
# LISTA KIT / COMPONENTI
# -------------------------------
@admin_required
def kit_list(request):
    """
    Vista di sola lettura di tutti i componenti kit,
    con filtro sul nome del kit (prodotto padre).
    """
    qs = (
        KitComponent.objects
        .select_related("kit")
        .all()
        .order_by("kit__name", "name")
    )

    kit_name = request.GET.get("kit")
    if kit_name:
        qs = qs.filter(kit__name__icontains=kit_name)

    context = {
        "kits": qs,
        "kit_name": kit_name or "",
    }
    return render(request, "backoffice/kit_list.html", context)
    
# -------------------------------
# LISTA PAGINE CMS
# -------------------------------
@admin_required
def cms_page_list(request):
    qs = Page.objects.all().order_by("title")

    title = request.GET.get("title")
    if title:
        qs = qs.filter(title__icontains=title)

    context = {
        "pages": qs,
        "title": title or "",
    }
    return render(request, "backoffice/cms_page_list.html", context)
    
# -------------------------------
# LISTA UTENTI COMPLETA
# -------------------------------
from accounts.models import User  # se non già importato

# -------------------------------
# RECENSIONI PRODOTTI (BACKOFFICE)
# -------------------------------
@content_staff_required
def review_list(request):
    """
    Lista recensioni con filtri base per:
    - prodotto
    - partner
    - utente
    - rating
    - stato approvazione
    """
    reviews = ProductRating.objects.select_related(
        "product",
        "product__supplier",
        "user"
    ).all()

    product_id = request.GET.get("product")
    partner_id = request.GET.get("partner")
    user_email = request.GET.get("user")
    rating = request.GET.get("rating")
    status = request.GET.get("status")

    if product_id:
        reviews = reviews.filter(product_id=product_id)

    if partner_id:
        reviews = reviews.filter(product__supplier_id=partner_id)

    if user_email:
        reviews = reviews.filter(user__email__icontains=user_email)

    if rating:
        reviews = reviews.filter(rating=rating)

    if status == "approved":
        reviews = reviews.filter(is_approved=True)
    elif status == "pending":
        reviews = reviews.filter(is_approved=False)

    reviews = reviews.order_by("-created_at")

    products = Product.objects.order_by("name")
    partners = PartnerProfile.objects.order_by("company_name")

    context = {
        "reviews": reviews,
        "products": products,
        "partners": partners,
        "product_id": int(product_id) if product_id else None,
        "partner_id": int(partner_id) if partner_id else None,
        "user_email": user_email or "",
        "rating": rating or "",
        "status": status or "",
    }
    return render(request, "backoffice/review_list.html", context)


@content_staff_required
def review_detail(request, review_id):
    """
    Dettaglio singola recensione, solo lettura + pulsanti moderazione.
    """
    review = get_object_or_404(
        ProductRating.objects.select_related("product", "product__supplier", "user"),
        id=review_id,
    )
    return render(request, "backoffice/review_detail.html", {"review": review})


@content_staff_required
def review_moderate(request, review_id):
    """
    Azione di moderazione (approva / rifiuta) dal backoffice.
    """
    review = get_object_or_404(ProductRating, id=review_id)

    action = request.POST.get("action")
    if action == "approve":
        review.is_approved = True
    elif action == "reject":
        review.is_approved = False

    # campi audit se li hai aggiunti nello STEP 2
    if hasattr(review, "moderated_by"):
        review.moderated_by = request.user
    if hasattr(review, "moderated_at"):
        from django.utils import timezone
        review.moderated_at = timezone.now()

    review.save()

    return HttpResponseRedirect(reverse("backoffice:review_list"))


@admin_required
def review_approve(request, pk):
    """
    Imposta la recensione come approvata e registra chi/ quando.
    """
    review = get_object_or_404(ProductRating, pk=pk)
    review.moderation_status = ProductRating.STATUS_APPROVED
    review.is_approved = True
    review.moderated_by = request.user
    review.moderated_at = timezone.now()
    review.save(update_fields=["moderation_status", "is_approved", "moderated_by", "moderated_at"])
    messages.success(
        request,
        f"Recensione #{review.id} per '{review.product}' approvata.",
    )
    return HttpResponseRedirect(reverse("backoffice:review_list"))


@admin_required
def review_reject(request, pk):
    """
    Imposta la recensione come rifiutata e registra chi/ quando.
    """
    review = get_object_or_404(ProductRating, pk=pk)
    review.moderation_status = ProductRating.STATUS_REJECTED
    review.is_approved = False
    review.moderated_by = request.user
    review.moderated_at = timezone.now()
    review.save(update_fields=["moderation_status", "is_approved", "moderated_by", "moderated_at"])
    messages.warning(
        request,
        f"Recensione #{review.id} per '{review.product}' segnata come RIFIUTATA.",
    )
    return HttpResponseRedirect(reverse("backoffice:review_list"))

@admin_required
def user_list(request):
    qs = User.objects.all().order_by("email")

    # FILTRI
    email = request.GET.get("email")
    role = request.GET.get("role")

    if email:
        qs = qs.filter(email__icontains=email)

    if role:
        qs = qs.filter(role=role)

    context = {
        "users": qs,
        "email": email or "",
        "role": role or "",
    }
    return render(request, "backoffice/user_list.html", context)
    
# -------------------------------
# COMMISSIONI PARTNER
# -------------------------------
@admin_required
def partner_commission_list(request):
    """
    Vista commissioni partner con:
    - total_revenue: somma delle righe ordine (order_items.total_price) nel periodo
    - portal_commissions: commissioni maturate lato portale (commission_amount nel periodo)
    - partner_earnings: quota maturata dai partner nel periodo
    - unliquidated_commissions: commissioni ancora da liquidare (quota portale)
    - liquidated_commissions: commissioni già liquidate (quota portale)
    """

    company = request.GET.get("company")
    email = request.GET.get("email")
    active = request.GET.get("active")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    # Ha effettuato una ricerca solo se almeno un filtro è valorizzato
    has_filters = any([company, email, active, start_date, end_date])

    # base queryset
    qs = PartnerProfile.objects.select_related("user")

    if has_filters:
        # filtri base (ragione sociale, email, attivo)
        if company:
            qs = qs.filter(company_name__icontains=company)

        if email:
            qs = qs.filter(user__email__icontains=email)

        if active == "yes":
            qs = qs.filter(is_active=True)
        elif active == "no":
            qs = qs.filter(is_active=False)

        # filtro periodo DA APPLICARE dentro le SUM
        # NB: le commissioni "maturano" solo su righe COMPLETED con commission_amount > 0.
        period_filter = Q(order_items__partner_status=OrderItem.PARTNER_STATUS_COMPLETED, order_items__commission_amount__gt=0)
        if start_date:
            period_filter &= Q(order_items__order__created_at__date__gte=start_date)
        if end_date:
            period_filter &= Q(order_items__order__created_at__date__lte=end_date)

        # annotazioni per il periodo selezionato
        qs = (
            qs.annotate(
                # FATTURATO del periodo (tutto, partner + portale)
                total_revenue=Coalesce(
                    Sum("order_items__total_price", filter=period_filter),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
            .annotate(
                # COMMISSIONE PORTALE = somma commission_amount nel periodo
                portal_commissions=Coalesce(
                    Sum(
                        "order_items__commission_amount",
                        filter=period_filter,
                    ),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
            .annotate(
                # IMPORTO PARTNER DA LIQUIDARE nel periodo
                partner_earnings=Coalesce(
                    Sum(
                        "order_items__partner_earnings",
                        filter=period_filter & Q(order_items__is_liquidated=False, order_items__payout__isnull=True),
                    ),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
            .annotate(
                # commissioni ancora da liquidare nel periodo (quota portale)
                unliquidated_commissions=Coalesce(
                    Sum(
                        "order_items__commission_amount",
                        filter=period_filter & Q(order_items__is_liquidated=False, order_items__payout__isnull=True),
                    ),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
            .annotate(
                # commissioni già liquidate nel periodo (quota portale)
                liquidated_commissions=Coalesce(
                    Sum(
                        "order_items__commission_amount",
                        filter=period_filter & Q(order_items__is_liquidated=True),
                    ),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            )
            .order_by("company_name")
        )

        # ✅ Mostra SOLO i partner che nel periodo hanno ancora qualcosa da liquidare
        qs = qs.filter(
            unliquidated_commissions__gt=Decimal("0.00"),
        )

        # Totali calcolati solo sui partner ancora visibili
        total_commissions = qs.aggregate(
            total=Coalesce(
                Sum("partner_earnings"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )["total"]

        total_to_liquidate = qs.aggregate(
            total=Coalesce(
                Sum("unliquidated_commissions"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )["total"]

        total_liquidated = qs.aggregate(
            total=Coalesce(
                Sum("liquidated_commissions"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )["total"]

    else:
        # nessun filtro: non mostriamo alcun partner/valore
        qs = PartnerProfile.objects.none()
        total_commissions = Decimal("0.00")
        total_to_liquidate = Decimal("0.00")
        total_liquidated = Decimal("0.00")

    context = {
        "partners": qs,
        "company": company or "",
        "email": email or "",
        "active": active or "",
        "period_start": period_start_str or "",
        "period_end": period_end_str or "",
        "total_commissions": total_commissions,
        "total_to_liquidate": total_to_liquidate,
        "total_liquidated": total_liquidated,
        "show_results": has_filters,  # controlla la visibilità delle card nel template
    }

    return render(request, "backoffice/partner_commission_list.html", context)

@admin_required
def commission_report(request):
    """
    Dashboard report commissioni:
    - KPI globali sul periodo
    - Grafico andamento giornaliero (fatturato / commissioni / guadagni partner)
    - Grafico top partner
    - Grafico fatturato & guadagni per mese
    - Grafico top categorie prodotto
    - Tabella riepilogo per partner
    - Grafico di confronto tra due partner selezionati
    """

    partner_id = request.GET.get("partner")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")
    compare_partner_1_id = request.GET.get("compare_partner_1")
    compare_partner_2_id = request.GET.get("compare_partner_2")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    # Base queryset: tutte le righe d'ordine con commissione (non annullate)
    items = (
        OrderItem.objects
        .select_related("order", "partner", "product", "partner__user", "product__category")
        .exclude(order__status="cancelled")
        .filter(commission_amount__gt=0)
    )

    if partner_id:
        items = items.filter(partner_id=partner_id)
    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)

    # ===========================
    #  KPI GLOBALI
    # ===========================
    agg = items.aggregate(
        total_revenue=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_portal_commission=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        order_count=Count("order", distinct=True),
        partner_count=Count("partner", distinct=True),
    )

    total_revenue = agg["total_revenue"] or Decimal("0.00")
    total_portal_commission = agg["total_portal_commission"] or Decimal("0.00")
    total_partner_earnings = total_revenue - total_portal_commission
    order_count = agg["order_count"] or 0
    partner_count = agg["partner_count"] or 0

    # ===========================
    #  RIEPILOGO PER PARTNER
    #  (tabella + grafico top partner)
    # ===========================
    partner_rows_qs = (
        items.values("partner__id", "partner__company_name", "partner__user__email")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            order_count=Count("order", distinct=True),
        )
        .order_by("-revenue")
    )

    partner_rows = []
    chart_partner_labels_list = []
    chart_partner_revenue_list = []
    chart_partner_partner_earnings_list = []

    for row in partner_rows_qs:
        revenue = row["revenue"] or Decimal("0.00")
        portal_commission = row["portal_commission"] or Decimal("0.00")
        partner_earnings = revenue - portal_commission

        row["partner_earnings"] = partner_earnings
        partner_rows.append(row)

        chart_partner_labels_list.append(row["partner__company_name"])
        chart_partner_revenue_list.append(float(revenue))
        chart_partner_partner_earnings_list.append(float(partner_earnings))

    # ===========================
    #  GRAFICO ANDAMENTO GIORNALIERO
    # ===========================
    time_qs = (
        items
        .annotate(day=TruncDate("order__created_at"))
        .values("day")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("day")
    )

    chart_time_labels_list = []
    chart_time_revenue_list = []
    chart_time_portal_commission_list = []
    chart_time_partner_earnings_list = []

    for row in time_qs:
        day = row["day"]
        revenue = row["revenue"] or Decimal("0.00")
        portal_commission = row["portal_commission"] or Decimal("0.00")
        partner_earnings = revenue - portal_commission

        chart_time_labels_list.append(day.strftime("%d/%m/%Y"))
        chart_time_revenue_list.append(float(revenue))
        chart_time_portal_commission_list.append(float(portal_commission))
        chart_time_partner_earnings_list.append(float(partner_earnings))

    # ===========================
    #  GRAFICO PER MESE
    # ===========================
    month_qs = (
        items
        .annotate(month=TruncMonth("order__created_at"))
        .values("month")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("month")
    )

    chart_month_labels_list = []
    chart_month_revenue_list = []
    chart_month_partner_earnings_list = []

    for row in month_qs:
        m = row["month"].strftime("%m/%Y")
        revenue = row["revenue"] or Decimal("0.00")
        comm = row["portal_commission"] or Decimal("0.00")
        partner_earn = revenue - comm

        chart_month_labels_list.append(m)
        chart_month_revenue_list.append(float(revenue))
        chart_month_partner_earnings_list.append(float(partner_earn))

    # ===========================
    #  GRAFICO TOP CATEGORIE
    # ===========================
    category_qs = (
        items.values("product__category__name")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("-revenue")
    )

    chart_category_labels_list = []
    chart_category_revenue_list = []
    chart_category_partner_earnings_list = []

    for row in category_qs:
        name = row["product__category__name"] or "Senza categoria"
        revenue = row["revenue"] or Decimal("0.00")
        comm = row["portal_commission"] or Decimal("0.00")
        partner_earn = revenue - comm

        chart_category_labels_list.append(name)
        chart_category_revenue_list.append(float(revenue))
        chart_category_partner_earnings_list.append(float(partner_earn))

    # ===========================
    #  GRAFICO CONFRONTO TRA DUE PARTNER
    # ===========================
    compare_metric_labels_list = []
    compare_partner1_values_list = []
    compare_partner2_values_list = []
    compare_partner_1_name = ""
    compare_partner_2_name = ""

    partners_qs = PartnerProfile.objects.order_by("company_name")

    if compare_partner_1_id and compare_partner_2_id and compare_partner_1_id != compare_partner_2_id:
        try:
            p1 = partners_qs.get(id=compare_partner_1_id)
            p2 = partners_qs.get(id=compare_partner_2_id)
        except PartnerProfile.DoesNotExist:
            p1 = p2 = None
        else:
            compare_partner_1_name = p1.company_name
            compare_partner_2_name = p2.company_name

            base_qs = items  # già filtrato per periodo / partner principale

            def agg_partner(qs):
                a = qs.aggregate(
                    revenue=Coalesce(
                        Sum("total_price"),
                        Value(Decimal("0.00")),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    ),
                    portal_commission=Coalesce(
                        Sum("commission_amount"),
                        Value(Decimal("0.00")),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    ),
                    order_count=Count("order", distinct=True),
                )
                revenue_ = a["revenue"] or Decimal("0.00")
                portal_commission_ = a["portal_commission"] or Decimal("0.00")
                partner_earnings_ = revenue_ - portal_commission_
                order_count_ = a["order_count"] or 0
                return revenue_, portal_commission_, partner_earnings_, order_count_

            rev1, comm1, earn1, orders1 = agg_partner(base_qs.filter(partner_id=compare_partner_1_id))
            rev2, comm2, earn2, orders2 = agg_partner(base_qs.filter(partner_id=compare_partner_2_id))

            compare_metric_labels_list = [
                "Fatturato",
                "Commissioni portale",
                "Guadagno partner",
                "N. ordini",
            ]
            compare_partner1_values_list = [
                float(rev1),
                float(comm1),
                float(earn1),
                int(orders1),
            ]
            compare_partner2_values_list = [
                float(rev2),
                float(comm2),
                float(earn2),
                int(orders2),
            ]

    # ===========================
    #  SERIALIZZAZIONE PER CHART.JS
    # ===========================
    chart_time_labels = json.dumps(chart_time_labels_list)
    chart_time_revenue = json.dumps(chart_time_revenue_list)
    chart_time_portal_commission = json.dumps(chart_time_portal_commission_list)
    chart_time_partner_earnings = json.dumps(chart_time_partner_earnings_list)

    chart_partner_labels = json.dumps(chart_partner_labels_list)
    chart_partner_revenue = json.dumps(chart_partner_revenue_list)
    chart_partner_partner_earnings = json.dumps(chart_partner_partner_earnings_list)

    chart_month_labels = json.dumps(chart_month_labels_list)
    chart_month_revenue = json.dumps(chart_month_revenue_list)
    chart_month_partner_earnings = json.dumps(chart_month_partner_earnings_list)

    chart_category_labels = json.dumps(chart_category_labels_list)
    chart_category_revenue = json.dumps(chart_category_revenue_list)
    chart_category_partner_earnings = json.dumps(chart_category_partner_earnings_list)

    chart_compare_metric_labels = json.dumps(compare_metric_labels_list)
    chart_compare_partner1_values = json.dumps(compare_partner1_values_list)
    chart_compare_partner2_values = json.dumps(compare_partner2_values_list)

    # ===========================
    #  CONTEXT
    # ===========================
    context = {
        "partners": partners_qs,
        "selected_partner": int(partner_id) if partner_id else None,
        "period_start": period_start_str or "",
        "period_end": period_end_str or "",
        "total_revenue": total_revenue,
        "total_portal_commission": total_portal_commission,
        "total_partner_earnings": total_partner_earnings,
        "order_count": order_count,
        "partner_count": partner_count,
        "partner_rows": partner_rows,

        "chart_time_labels": chart_time_labels,
        "chart_time_revenue": chart_time_revenue,
        "chart_time_portal_commission": chart_time_portal_commission,
        "chart_time_partner_earnings": chart_time_partner_earnings,

        "chart_partner_labels": chart_partner_labels,
        "chart_partner_revenue": chart_partner_revenue,
        "chart_partner_partner_earnings": chart_partner_partner_earnings,

        "chart_month_labels": chart_month_labels,
        "chart_month_revenue": chart_month_revenue,
        "chart_month_partner_earnings": chart_month_partner_earnings,

        "chart_category_labels": chart_category_labels,
        "chart_category_revenue": chart_category_revenue,
        "chart_category_partner_earnings": chart_category_partner_earnings,

        "chart_compare_metric_labels": chart_compare_metric_labels,
        "chart_compare_partner1_values": chart_compare_partner1_values,
        "chart_compare_partner2_values": chart_compare_partner2_values,
        "compare_partner_1_name": compare_partner_1_name,
        "compare_partner_2_name": compare_partner_2_name,
        "selected_compare_partner_1": int(compare_partner_1_id) if compare_partner_1_id else None,
        "selected_compare_partner_2": int(compare_partner_2_id) if compare_partner_2_id else None,
    }

    return render(request, "backoffice/commission_report.html", context)


@admin_required
def commission_report_detail(request):
    """
    Drill-down: elenco di tutte le righe d'ordine che compongono le commissioni
    del report, con filtri avanzati:
    - partner
    - periodo (da/a)
    - categoria prodotto
    - struttura
    - stato ordine
    - stato payout (liquidata / non liquidata)
    """

    partner_id = request.GET.get("partner")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")
    category_id = request.GET.get("category")
    structure_id = request.GET.get("structure")
    order_status = request.GET.get("order_status")
    payout_status = request.GET.get("payout_status")  # '', 'liquidated', 'unliquidated'

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    items = (
        OrderItem.objects
        .select_related(
            "order",
            "order__structure",
            "product",
            "product__category",
            "partner",
        )
        .exclude(order__status=Order.STATUS_CANCELLED)
        .filter(commission_amount__gt=0)
    )

    if partner_id:
        items = items.filter(partner_id=partner_id)
    if category_id:
        items = items.filter(product__category_id=category_id)
    if structure_id:
        items = items.filter(order__structure_id=structure_id)
    if order_status:
        items = items.filter(order__status=order_status)
    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)

    if payout_status == "liquidated":
        items = items.filter(is_liquidated=True)
    elif payout_status == "unliquidated":
        items = items.filter(is_liquidated=False)

    items = items.order_by("-order__created_at")

    # KPI di riepilogo
    agg = items.aggregate(
        total_revenue=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_portal_commission=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
    )

    total_revenue = agg["total_revenue"] or Decimal("0.00")
    total_portal_commission = agg["total_portal_commission"] or Decimal("0.00")
    total_partner_earnings = total_revenue - total_portal_commission

    # calcolo guadagno partner per ogni riga (per il template)
    for it in items:
        rev = it.total_price or Decimal("0.00")
        comm = it.commission_amount or Decimal("0.00")
        it.calculated_partner_earnings = rev - comm

    partners = PartnerProfile.objects.order_by("company_name")
    categories = Category.objects.order_by("name")
    structures = ClientStructure.objects.order_by("id")  # ordina per id, sicuro

    context = {
        "items": items,
        "partners": partners,
        "categories": categories,
        "structures": structures,
        "selected_partner": int(partner_id) if partner_id else None,
        "selected_category": int(category_id) if category_id else None,
        "selected_structure": int(structure_id) if structure_id else None,
        "selected_order_status": order_status or "",
        "selected_payout_status": payout_status or "",
        "period_start": period_start_str or "",
        "period_end": period_end_str or "",
        "total_revenue": total_revenue,
        "total_portal_commission": total_portal_commission,
        "total_partner_earnings": total_partner_earnings,
        "order_status_choices": Order.STATUS_CHOICES,
    }

    return render(request, "backoffice/commission_report_detail.html", context)


@admin_required
def commission_report_export_csv(request):
    """
    Export CSV del report commissioni (aggregato per partner),
    con gli stessi filtri di commission_report.
    """

    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")
    partner_id = request.GET.get("partner")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    items = (
        OrderItem.objects
        .select_related("order", "partner", "partner__user")
        .exclude(order__status="cancelled")
        .filter(commission_amount__gt=0)
    )

    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)
    if partner_id:
        items = items.filter(partner_id=partner_id)

    partner_rows_qs = (
        items.filter(partner__isnull=False)
        .values("partner__id", "partner__company_name", "partner__user__email")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("partner__company_name")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="report_commissioni_portale.csv"'

    writer = csv.writer(response, delimiter=";")

    writer.writerow([
        "Partner",
        "Email",
        "Fatturato (€)",
        "Commissioni portale (€)",
        "Guadagno partner (€)",
        "Periodo da",
        "Periodo a",
    ])

    for row in partner_rows_qs:
        rev = row["revenue"] or Decimal("0.00")
        comm = row["portal_commission"] or Decimal("0.00")
        partner_earn = rev - comm

        writer.writerow([
            row["partner__company_name"],
            row["partner__user__email"],
            f"{rev:.2f}",
            f"{comm:.2f}",
            f"{partner_earn:.2f}",
            period_start_str or "",
            period_end_str or "",
        ])

    return response

@admin_required
def commission_report_export_xlsx(request):
    """
    Export Excel (.xlsx) del report commissioni.
    - Sheet 1: riepilogo per partner
    - Sheet 2: dettaglio righe ordine
    """

    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")
    partner_id = request.GET.get("partner")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    items = (
        OrderItem.objects
        .select_related("order", "partner", "product", "partner__user", "product__category")
        .exclude(order__status="cancelled")
        .filter(commission_amount__gt=0)
    )

    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)
    if partner_id:
        items = items.filter(partner_id=partner_id)

    # ================
    #  CREA EXCEL
    # ================
    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo partner"

    # =====================
    #  FOGLIO 1: RIEPILOGO
    # =====================
    headers = [
        "Partner",
        "Email",
        "Fatturato (€)",
        "Commissioni portale (€)",
        "Guadagno partner (€)",
        "Periodo da",
        "Periodo a",
    ]

    # Header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    partner_rows_qs = (
        items.filter(partner__isnull=False)
        .values("partner__id", "partner__company_name", "partner__user__email")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("partner__company_name")
    )

    row_idx = 2
    for row in partner_rows_qs:
        rev = row["revenue"] or Decimal("0.00")
        comm = row["portal_commission"] or Decimal("0.00")
        partner_earn = rev - comm

        ws.cell(row=row_idx, column=1, value=row["partner__company_name"])
        ws.cell(row=row_idx, column=2, value=row["partner__user__email"])
        ws.cell(row=row_idx, column=3, value=float(rev))
        ws.cell(row=row_idx, column=4, value=float(comm))
        ws.cell(row=row_idx, column=5, value=float(partner_earn))
        ws.cell(row=row_idx, column=6, value=period_start_str or "")
        ws.cell(row=row_idx, column=7, value=period_end_str or "")

        row_idx += 1

    # Auto width foglio 1
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # ==========================
    #  FOGLIO 2: DETTAGLIO RIGHE
    # ==========================
    ws2 = wb.create_sheet("Dettaglio righe")

    headers2 = [
        "Data ordine",
        "ID ordine",
        "Cliente",
        "Partner",
        "Prodotto",
        "Categoria",
        "Quantità",
        "Prezzo unitario (€)",
        "Totale riga (€)",
        "Commissione portale (€)",
        "Guadagno partner (€)",
        "Stato ordine",
    ]

    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    row_idx = 2
    for item in items:
        rev = item.total_price or Decimal("0.00")
        comm = item.commission_amount or Decimal("0.00")
        partner_earn = rev - comm

        ws2.cell(row=row_idx, column=1, value=item.order.created_at.strftime("%d/%m/%Y"))
        ws2.cell(row=row_idx, column=2, value=item.order.id)
        ws2.cell(row=row_idx, column=3, value="")
        ws2.cell(row=row_idx, column=4, value=item.partner.company_name if item.partner else "")
        ws2.cell(row=row_idx, column=5, value=item.product.name if item.product else "")
        ws2.cell(row=row_idx, column=6, value=item.product.category.name if item.product and item.product.category else "")
        ws2.cell(row=row_idx, column=7, value=item.quantity)
        ws2.cell(row=row_idx, column=8, value=float(item.unit_price))
        ws2.cell(row=row_idx, column=9, value=float(rev))
        ws2.cell(row=row_idx, column=10, value=float(comm))
        ws2.cell(row=row_idx, column=11, value=float(partner_earn))
        ws2.cell(row=row_idx, column=12, value=item.order.status)

        row_idx += 1

    # Auto width foglio 2
    for col_idx in range(1, len(headers2) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws2[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_length + 2

    # ============
    #  EXPORT
    # ============
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="report_commissioni_portale.xlsx"'
    wb.save(response)
    return response

@admin_required
def commission_report_export_pdf(request):
    """
    Export PDF del report commissioni:
    - riepilogo KPI globali
    - tabella riepilogo per partner (come nel report principale)
    """

    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")
    partner_id = request.GET.get("partner")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    # Query di base (stessa logica del report)
    items = (
        OrderItem.objects
        .select_related("order", "partner", "product", "partner__user")
        .exclude(order__status="cancelled")
        .filter(commission_amount__gt=0)
    )

    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)
    if partner_id:
        items = items.filter(partner_id=partner_id)

    from django.db.models import Count

    agg = items.aggregate(
        total_revenue=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_portal_commission=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        order_count=Count("order", distinct=True),
        partner_count=Count("partner", distinct=True),
    )

    total_revenue = agg["total_revenue"] or Decimal("0.00")
    total_portal_commission = agg["total_portal_commission"] or Decimal("0.00")
    total_partner_earnings = (total_revenue - total_portal_commission).quantize(Decimal("0.01"))
    order_count = agg["order_count"] or 0
    partner_count = agg["partner_count"] or 0

    # Riepilogo per partner (come nella tabella principale)
    partner_rows_qs = (
        items.filter(partner__isnull=False)
        .values("partner__id", "partner__company_name", "partner__user__email")
        .annotate(
            revenue=Coalesce(
                Sum("total_price"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            portal_commission=Coalesce(
                Sum("commission_amount"),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by("partner__company_name")
    )

    partner_rows = []
    for row in partner_rows_qs:
        rev = row["revenue"] or Decimal("0.00")
        comm = row["portal_commission"] or Decimal("0.00")
        partner_earn = rev - comm
        row["partner_earnings"] = partner_earn
        partner_rows.append(row)

    # ==================
    #  CREAZIONE PDF
    # ==================
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="report_commissioni_portale.pdf"'

    c = NumberedCanvas(response, pagesize=A4)
    width, height = A4

    margin_x = 20 * mm
    margin_top = height - 20 * mm
    line_height = 6 * mm

    y = margin_top

    def draw_header():
        nonlocal y
        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin_x, y, "Report commissioni portale & guadagni partner")
        y -= line_height

        c.setFont("Helvetica", 9)
        period_text = "Periodo: "
        if period_start_str or period_end_str:
            period_text += f"{period_start_str or 'inizio'} → {period_end_str or 'oggi'}"
        else:
            period_text += "completo"
        c.drawString(margin_x, y, period_text)
        y -= line_height * 1.5

    def draw_kpi():
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin_x, y, f"Fatturato totale: {total_revenue:.2f} €")
        y -= line_height
        c.drawString(margin_x, y, f"Commissioni portale: {total_portal_commission:.2f} €")
        y -= line_height
        c.drawString(margin_x, y, f"Guadagni partner: {total_partner_earnings:.2f} €")
        y -= line_height
        c.drawString(margin_x, y, f"Ordini: {order_count}  ·  Partner: {partner_count}")
        y -= line_height * 1.5

    def new_page():
        nonlocal y
        c.showPage()
        y = margin_top
        draw_header()
        y -= line_height * 0.5

    # Disegna header + KPI sulla prima pagina
    draw_header()
    draw_kpi()

    # ==================
    #  TABELLA PARTNER
    # ==================
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin_x, y, "Dettaglio commissioni per partner")
    y -= line_height * 1.2

    # Header tabella
    c.setFont("Helvetica-Bold", 8)
    col_partner = margin_x
    col_email = margin_x + 60 * mm
    col_rev = margin_x + 115 * mm
    col_comm = margin_x + 145 * mm
    col_earn = margin_x + 175 * mm

    c.drawString(col_partner, y, "Partner")
    c.drawString(col_email, y, "Email")
    c.drawRightString(col_rev + 20, y, "Fatturato (€)")
    c.drawRightString(col_comm + 20, y, "Comm. portale (€)")
    c.drawRightString(col_earn + 20, y, "Guadagno partner (€)")
    y -= line_height

    c.setFont("Helvetica", 8)

    if not partner_rows:
        c.drawString(margin_x, y, "Nessun dato disponibile per i filtri selezionati.")
    else:
        for row in partner_rows:
            if y < 30 * mm:
                new_page()
                c.setFont("Helvetica-Bold", 8)
                c.drawString(col_partner, y, "Partner")
                c.drawString(col_email, y, "Email")
                c.drawRightString(col_rev + 20, y, "Fatturato (€)")
                c.drawRightString(col_comm + 20, y, "Comm. portale (€)")
                c.drawRightString(col_earn + 20, y, "Guadagno partner (€)")
                y -= line_height
                c.setFont("Helvetica", 8)

            partner_name = row["partner__company_name"] or ""
            email = row["partner__user__email"] or ""
            rev = row["revenue"] or Decimal("0.00")
            comm = row["portal_commission"] or Decimal("0.00")
            partner_earn = row["partner_earnings"] or Decimal("0.00")

            # testo (tagliato se troppo lungo)
            if len(partner_name) > 40:
                partner_name = partner_name[:37] + "..."
            if len(email) > 40:
                email = email[:37] + "..."

            c.drawString(col_partner, y, partner_name)
            c.drawString(col_email, y, email)
            c.drawRightString(col_rev + 20, y, f"{rev:.2f}")
            c.drawRightString(col_comm + 20, y, f"{comm:.2f}")
            c.drawRightString(col_earn + 20, y, f"{partner_earn:.2f}")
            y -= line_height

    c.showPage()
    c.save()
    return response

@admin_required
def commission_partner_pdf(request, partner_id):
    """
    PDF di dettaglio per SINGOLO partner:
    - header istituzionale con logo e dati azienda
    - riepilogo KPI per quel partner
    - lista completa delle righe d’ordine filtrate per periodo
    """

    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    # Partner
    partner = get_object_or_404(PartnerProfile, id=partner_id)

    # Righe ordine del partner
    items = (
        OrderItem.objects
        .select_related(
            "order",
            "product",
            "partner",
            "product__category",
        )
        .filter(partner=partner)
        .exclude(order__status="cancelled")
    )

    if start_date:
        items = items.filter(order__created_at__date__gte=start_date)
    if end_date:
        items = items.filter(order__created_at__date__lte=end_date)

    items = items.order_by("order__created_at")

    # KPI per questo partner
    agg = items.aggregate(
        revenue=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        portal_comm=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        order_count=Count("order", distinct=True),
    )

    revenue = agg["revenue"] or Decimal("0.00")
    portal_comm = agg["portal_comm"] or Decimal("0.00")
    partner_earn = revenue - portal_comm
    order_count = agg["order_count"] or 0

    # ================
    #  CREA PDF
    # ================
    response = HttpResponse(content_type="application/pdf")
    safe_name = (partner.company_name or "partner").replace(" ", "_")
    filename = f"report_partner_{safe_name}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Usa il canvas numerato con footer istituzionale
    c = NumberedCanvas(response, pagesize=A4)
    width, height = A4

    margin_x = 20 * mm
    margin_top = height - 25 * mm
    line_h = 6 * mm

    y = margin_top

    def draw_brand_header():
        """
        Header istituzionale con:
        - logo
        - dati azienda
        - titolo report
        - periodo
        - riga bianca e linea di separazione
        """
        nonlocal y

        # "Logo" stilizzato
        brand_y = height - 18 * mm
        c.setFillColor(colors.HexColor("#2563EB"))
        c.roundRect(margin_x, brand_y - 6 * mm, 12 * mm, 10 * mm, 2 * mm, fill=1, stroke=0)

        # Nome del portale
        c.setFillColor(colors.HexColor("#0F172A"))
        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin_x + 15 * mm, brand_y, "B2B Turismo")

        # Dati azienda
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#4B5563"))
        c.drawRightString(width - margin_x, brand_y, "B2B Turismo S.r.l.")
        c.drawRightString(width - margin_x, brand_y - 4 * mm, "Via Roma 123, 80100 Napoli (NA)")
        c.drawRightString(width - margin_x, brand_y - 8 * mm, "info@b2bturismo.it – +39 081 000000")

        c.setFillColor(colors.black)

        # Titolo report
        y_local = brand_y - 16 * mm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin_x, y_local, f"Report commissioni — {partner.company_name}")
        y_local -= line_h

        # Periodo
        c.setFont("Helvetica", 9)
        period_text = "Periodo: "
        if period_start_str or period_end_str:
            period_text += f"{period_start_str or 'inizio'} → {period_end_str or 'oggi'}"
        else:
            period_text += "completo"
        c.drawString(margin_x, y_local, period_text)
        y_local -= line_h

        # Spazio extra sotto l’intestazione
        y_local -= 4 * mm

        # Linea orizzontale di separazione
        c.setStrokeColor(colors.HexColor("#CBD5E1"))  # grigio chiaro
        c.setLineWidth(0.8)
        c.line(margin_x, y_local, width - margin_x, y_local)
        c.setStrokeColor(colors.black)

        # Aggiorna Y per il contenuto sottostante
        y_local -= 8 * mm
        y = y_local

    def draw_kpi_block():
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin_x, y, f"Fatturato: {revenue:.2f} €")
        y -= line_h
        c.drawString(margin_x, y, f"Commissioni portale: {portal_comm:.2f} €")
        y -= line_h
        c.drawString(margin_x, y, f"Guadagno partner: {partner_earn:.2f} €")
        y -= line_h
        c.drawString(margin_x, y, f"Ordini: {order_count}")
        y -= line_h * 1.5

    def new_page():
        nonlocal y
        c.showPage()
        y = margin_top
        draw_brand_header()
        draw_kpi_block()

    # Prima pagina: header + KPI
    draw_brand_header()
    draw_kpi_block()

    # ================
    #  TABELLA DETTAGLIO
    # ================
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin_x, y, "Dettaglio righe ordine")
    y -= line_h * 1.2

    # Header tabella
    c.setFont("Helvetica-Bold", 8)
    col_date = margin_x
    col_order = margin_x + 25 * mm
    col_prod = margin_x + 50 * mm
    col_total = margin_x + 120 * mm
    col_comm = margin_x + 145 * mm
    col_earn = margin_x + 170 * mm

    c.drawString(col_date, y, "Data")
    c.drawString(col_order, y, "Ordine")
    c.drawString(col_prod, y, "Prodotto")
    c.drawRightString(col_total + 10, y, "Totale €")
    c.drawRightString(col_comm + 10, y, "Comm. €")
    c.drawRightString(col_earn + 10, y, "Guadagno €")
    y -= line_h

    c.setFont("Helvetica", 8)

    if not items.exists():
        c.drawString(margin_x, y, "Nessuna riga trovata per i filtri selezionati.")
    else:
        for item in items:
            if y < 30 * mm:
                new_page()
                # header tabella su nuova pagina
                c.setFont("Helvetica-Bold", 8)
                c.drawString(col_date, y, "Data")
                c.drawString(col_order, y, "Ordine")
                c.drawString(col_prod, y, "Prodotto")
                c.drawRightString(col_total + 10, y, "Totale €")
                c.drawRightString(col_comm + 10, y, "Comm. €")
                c.drawRightString(col_earn + 10, y, "Guadagno €")
                y -= line_h
                c.setFont("Helvetica", 8)

            d_str = item.order.created_at.strftime("%d/%m/%Y")
            prod_name = item.product.name if item.product else ""
            if len(prod_name) > 40:
                prod_name = prod_name[:37] + "..."

            rev = item.total_price or Decimal("0.00")
            comm = item.commission_amount or Decimal("0.00")
            earn = rev - comm

            c.drawString(col_date, y, d_str)
            c.drawString(col_order, y, str(item.order.id))
            c.drawString(col_prod, y, prod_name)
            c.drawRightString(col_total + 10, y, f"{rev:.2f}")
            c.drawRightString(col_comm + 10, y, f"{comm:.2f}")
            c.drawRightString(col_earn + 10, y, f"{earn:.2f}")

            y -= line_h

    c.showPage()
    c.save()
    return response



@admin_required
def partner_commission_export_csv(request):
    """
    Esporta in CSV il riepilogo commissioni per partner
    usando gli stessi filtri (compreso periodo) della lista.
    Qui per 'Commissioni maturate' intendiamo la quota PORTALE (commission_amount).
    """
    company = request.GET.get("company")
    email = request.GET.get("email")
    active = request.GET.get("active")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    qs = PartnerProfile.objects.select_related("user")

    if start_date:
        qs = qs.filter(order_items__order__created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(order_items__order__created_at__date__lte=end_date)

    if company:
        qs = qs.filter(company_name__icontains=company)

    if email:
        qs = qs.filter(user__email__icontains=email)

    if active == "yes":
        qs = qs.filter(is_active=True)
    elif active == "no":
        qs = qs.filter(is_active=False)

    # filtro periodo anche per l'export
    period_filter = Q()
    if start_date:
        period_filter &= Q(order_items__order__created_at__date__gte=start_date)
    if end_date:
        period_filter &= Q(order_items__order__created_at__date__lte=end_date)

    qs = (
        qs.annotate(
            total_revenue=Coalesce(
                Sum("order_items__total_price", filter=period_filter),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .annotate(
            # commissioni "maturate" lato portale
            matured_commissions=Coalesce(
                Sum("order_items__commission_amount", filter=period_filter),
                Value(Decimal("0.00")),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .order_by("company_name")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="commissioni_partner.csv"'

    writer = csv.writer(response, delimiter=";")

    writer.writerow([
        "Partner",
        "Email",
        "Attivo",
        "Percentuale commissione",
        "Fatturato totale",
        "Commissioni portale maturate",
        "Periodo da",
        "Periodo a",
    ])

    for p in qs:
        attivo = "Sì" if p.is_active else "No"
        percent = p.default_commission_percent or Decimal("0.00")
        total_revenue = getattr(p, "total_revenue", Decimal("0.00")) or Decimal("0.00")
        commission_value = getattr(p, "matured_commissions", Decimal("0.00")) or Decimal("0.00")

        writer.writerow([
            p.company_name,
            p.user.email,
            attivo,
            f"{percent:.2f}",
            f"{total_revenue:.2f}",
            f"{commission_value:.2f}",
            period_start_str or "",
            period_end_str or "",
        ])

    return response
 
@admin_required
def partner_payout_report(request, payout_id):
    """
    Versione 'report' del payout, pensata per la stampa / PDF.
    Layout minimale stile report commissioni.
    """
    payout = get_object_or_404(
        PartnerPayout.objects.select_related("partner", "partner__user"),
        id=payout_id,
    )

    # Usiamo le righe collegate a questo payout
    items_qs = (
        payout.items
        .select_related("order", "order__client", "product")
        .order_by("order__created_at", "id")
    )

    aggregates = items_qs.aggregate(
        total_row_amount=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_commission=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_partner_net=Coalesce(
            Sum("partner_earnings"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
    )

    context = {
        "payout": payout,
        "items": items_qs,
        "total_row_amount": aggregates["total_row_amount"],
        "total_commission": aggregates["total_commission"],
        "total_partner_net": aggregates["total_partner_net"],
    }
    return render(request, "backoffice/partner_payout_report.html", context)
 
    
@admin_required
def partner_payout_create(request, partner_id):
    """
    Crea o aggiorna un PartnerPayout per il partner e il periodo indicati.

    Regole:
    - Se esiste un payout in stato BOZZA per (partner, periodo) -> aggiunge le nuove righe
      non liquidate al suo totale.
    - Se NON esiste un payout in BOZZA (magari ce ne sono di già pagati) -> crea un
      nuovo payout in stato BOZZA solo con le nuove righe.

    In tutti i casi, le righe vengono associate al payout (campo OrderItem.payout).
    La liquidazione (is_liquidated=True) avviene SOLO quando il payout passa a "Pagato".
    """

    partner = get_object_or_404(PartnerProfile, id=partner_id)

    if request.method != "POST":
        return HttpResponseRedirect(reverse("backoffice:partner_commission_list"))

    # Filtri arrivati dal form nascosto (per tornare alla stessa vista filtrata)
    company = request.POST.get("company")
    email = request.POST.get("email")
    active = request.POST.get("active")
    period_start_str = request.POST.get("period_start")
    period_end_str = request.POST.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    # Costruiamo SUBITO l'URL di redirect con gli stessi filtri
    base_url = reverse("backoffice:partner_commission_list")
    params = {}
    for name, value in [
        ("company", company),
        ("email", email),
        ("active", active),
        ("period_start", period_start_str),
        ("period_end", period_end_str),
    ]:
        if value:
            params[name] = value

    redirect_url = f"{base_url}?{urlencode(params)}" if params else base_url

    if not start_date or not end_date:
        messages.error(
            request,
            "Devi specificare un periodo valido (dal/al) per creare un payout.",
        )
        return HttpResponseRedirect(redirect_url)

    # Righe d'ordine del periodo, NON ancora liquidate
    items_qs = OrderItem.objects.filter(
        partner=partner,
        order__created_at__date__gte=start_date,
        order__created_at__date__lte=end_date,
        is_liquidated=False,
        payout__isnull=True,  # evita doppio inserimento in payout già creati (bozze incluse)
    )

    if not items_qs.exists():
        messages.warning(
            request,
            "Non ci sono righe da liquidare nel periodo selezionato per questo partner "
            "oppure sono già state tutte liquidate.",
        )
        return HttpResponseRedirect(redirect_url)

    # Assicuriamoci che commission_amount e partner_earnings siano valorizzati
    for item in items_qs:
        if (
            item.commission_amount is None
            or item.partner_earnings is None
            or item.commission_amount == 0
            or item.partner_earnings == 0
        ):
            rate = (
                item.commission_rate
                or partner.default_commission_percent
                or Decimal("0.00")
            )
            item.commission_rate = rate
            item.calculate_commission(default_rate=rate)
            item.save(
                update_fields=[
                    "commission_rate",
                    "commission_amount",
                    "partner_earnings",
                ]
            )

    # Totale da liquidare AL PARTNER per queste NUOVE righe
    new_partner_earnings = items_qs.aggregate(
        total=Coalesce(
            Sum("partner_earnings"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )["total"]

    # Cerchiamo un payout ESISTENTE per questo partner/periodo in BOZZA
    existing_draft = PartnerPayout.objects.filter(
        partner=partner,
        period_start=start_date,
        period_end=end_date,
        status=PartnerPayout.STATUS_DRAFT,
    ).first()

    # Cerchiamo qualunque payout (bozza o pagato) solo per messaggio informativo
    existing_any = PartnerPayout.objects.filter(
        partner=partner,
        period_start=start_date,
        period_end=end_date,
    ).exclude(id=getattr(existing_draft, "id", None)).first()

    if existing_draft:
        # ➤ Caso 1 — aggiorno la bozza
        previous_total = existing_draft.total_commission or Decimal("0.00")
        existing_draft.total_commission = previous_total + new_partner_earnings
        existing_draft.save(update_fields=["total_commission"])

        payout = existing_draft

        messages.info(
            request,
            f"Payout in bozza già esistente per questo partner e periodo: "
            f"aggiunti {new_partner_earnings:.2f} € (nuove righe). "
            f"Totale aggiornato a {payout.total_commission:.2f} €.",
        )
    else:
        # ➤ Caso 2/3 — non esiste bozza: creo SEMPRE un nuovo payout
        payout = PartnerPayout.objects.create(
            partner=partner,
            period_start=start_date,
            period_end=end_date,
            total_commission=new_partner_earnings,
            status=PartnerPayout.STATUS_DRAFT,
        )

        if existing_any:
            messages.warning(
                request,
                f"Esisteva già un payout per questo periodo ma non era in bozza "
                f"(stato: {existing_any.get_status_display()}). "
                f"Ne è stato creato uno nuovo in stato 'Bozza'.",
            )
        else:
            messages.success(
                request,
                f"Creato nuovo payout in stato 'Bozza' per {partner.company_name} "
                f"({start_date} → {end_date}) per {new_partner_earnings:.2f} € (importo partner).",
            )

    # Le righe vengono associate al payout (la liquidazione avviene SOLO a payout=Pagato)
    items_qs.update(payout=payout)

    return HttpResponseRedirect(redirect_url)
    
    
@admin_required
def partner_payout_list(request):
    """
    Elenco dei PartnerPayout con filtri base (partner, stato),
    con numerazione batch per partner + periodo.
    """
    partner_id = request.GET.get("partner")
    status = request.GET.get("status")

    qs = PartnerPayout.objects.select_related("partner", "partner__user")

    if partner_id:
        qs = qs.filter(partner_id=partner_id)

    if status:
        qs = qs.filter(status=status)

    # Annotiamo il numero di batch per (partner, periodo_start, periodo_end)
    qs = qs.annotate(
        batch_number=Window(
            expression=RowNumber(),
            partition_by=[F("partner_id"), F("period_start"), F("period_end")],
            order_by=F("created_at").asc(),
        )
    ).order_by("-period_end", "-created_at")

    # partner per la tendina filtri
    partners = PartnerProfile.objects.order_by("company_name")

    # totale da pagare (solo confirmed, per esempio)
    total_confirmed = qs.filter(status=PartnerPayout.STATUS_CONFIRMED).aggregate(
        total=Coalesce(
            Sum("total_commission"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )["total"]

    context = {
        "payouts": qs,
        "partners": partners,
        "selected_partner": int(partner_id) if partner_id else None,
        "selected_status": status or "",
        "total_confirmed": total_confirmed,
    }

    return render(request, "backoffice/partner_payout_list.html", context)
    
@admin_required
def partner_payout_detail(request, payout_id):
    """
    Dettaglio di un PartnerPayout con elenco righe ordine coinvolte.
    La pagina è pensata anche per essere stampata / salvata in PDF dal browser.
    """
    payout = get_object_or_404(
        PartnerPayout.objects.select_related("partner", "partner__user"),
        id=payout_id,
    )

    # Righe d'ordine incluse in QUESTO payout (legate via OrderItem.payout)
    items_qs = (
        payout.items
        .select_related("order", "order__client", "product")
        .order_by("-order__created_at", "id")
    )

    # Totali di comodo (calcolati solo sulle righe di questo payout)
    aggregates = items_qs.aggregate(
        total_row_amount=Coalesce(
            Sum("total_price"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_commission=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        total_partner_net=Coalesce(
            Sum("partner_earnings"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
    )

    context = {
        "payout": payout,
        "items": items_qs,
        "total_row_amount": aggregates["total_row_amount"],
        "total_commission": aggregates["total_commission"],
        "total_partner_net": aggregates["total_partner_net"],
    }
    return render(request, "backoffice/partner_payout_detail.html", context)
 
 
@admin_required
def unliquidated_commission_list(request):
    """
    Elenco di tutte le righe d'ordine con commissioni non ancora liquidate.
    Utile per contabilità.
    """
    partner_id = request.GET.get("partner")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    qs = (
        OrderItem.objects
        .select_related("order", "product", "partner", "partner__user")
        .filter(
            commission_amount__gt=0,
            partner_status=OrderItem.PARTNER_STATUS_COMPLETED,
            is_liquidated=False,
            payout__isnull=True,
        )
    )

    if partner_id:
        qs = qs.filter(partner_id=partner_id)

    if start_date:
        qs = qs.filter(order__created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(order__created_at__date__lte=end_date)

    qs = qs.order_by("partner__company_name", "-order__created_at")

    total_unliquidated = qs.aggregate(
        total=Coalesce(
            Sum("commission_amount"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )["total"]

    partners = PartnerProfile.objects.order_by("company_name")

    context = {
        "items": qs,
        "partners": partners,
        "selected_partner": int(partner_id) if partner_id else None,
        "period_start": period_start_str or "",
        "period_end": period_end_str or "",
        "total_unliquidated": total_unliquidated,
    }

    return render(request, "backoffice/unliquidated_commission_list.html", context)
    
    
@admin_required
def liquidated_commission_list(request):
    """
    Elenco delle righe d'ordine con commissioni già liquidate.
    Filtrabile per partner e periodo, con link al payout (se individuabile).
    """
    partner_id = request.GET.get("partner")
    period_start_str = request.GET.get("period_start")
    period_end_str = request.GET.get("period_end")

    start_date = parse_date(period_start_str) if period_start_str else None
    end_date = parse_date(period_end_str) if period_end_str else None

    qs = (
        OrderItem.objects
        .select_related("order", "product", "partner", "partner__user")
        .filter(
            commission_amount__gt=0,
            is_liquidated=True,
        )
    )

    if partner_id:
        qs = qs.filter(partner_id=partner_id)

    if start_date:
        qs = qs.filter(order__created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(order__created_at__date__lte=end_date)

    # Proviamo ad agganciare il payout relativo:
    # cerchiamo un PartnerPayout per lo stesso partner e un periodo che includa la data ordine
    # Subquery: cerchiamo un payout che copra la data dell'ordine
    payout_subquery = PartnerPayout.objects.filter(
        partner=OuterRef("partner"),
        period_start__lte=OuterRef("order__created_at"),
        period_end__gte=OuterRef("order__created_at"),
    ).order_by("-period_end").values("id")[:1]

    qs = qs.annotate(payout_id=Subquery(payout_subquery))

    qs = qs.order_by("partner__company_name", "-order__created_at")

    total_liquidated = qs.aggregate(
        total=Coalesce(
            Sum("partner_earnings"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )["total"]

    partners = PartnerProfile.objects.order_by("company_name")

    context = {
        "items": qs,
        "partners": partners,
        "selected_partner": int(partner_id) if partner_id else None,
        "period_start": period_start_str or "",
        "period_end": period_end_str or "",
        "total_liquidated": total_liquidated,
    }

    return render(request, "backoffice/liquidated_commission_list.html", context)
